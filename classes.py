from dataclasses import dataclass
from typing import List
from datetime import date
from openpyxl import Workbook, load_workbook

@dataclass
class Student:
    fio: str
    id: int

    def __init__(self, new_fio: str, new_id: int):
        self.fio = new_fio
        self.id = new_id

@dataclass
class Specialization:
    name: str

    def __init__(self, new_name: str):
        self.name = new_name

@dataclass
class Subject:
    id: str
    name: str
    semester: int
    hours: int
    spec: Specialization

    def __init__(self, new_id: str, new_name: str, new_semester: int, new_hours: int, new_spec: Specialization):
        self.id = new_id
        self.name = new_name
        self.semester = new_semester
        self.hours = new_hours
        self.spec = new_spec

    def getSubject(path, compare_name):
        subject_wb = load_workbook(filename=path)
        ws = subject_wb.active

        for row in range(2, ws.max_row + 1):
            if (ws[row][1].value == compare_name):
                id = str(ws[row][0].value)
                name = str(ws[row][1].value)
                spec = Specialization(ws[row][2].value)
                semester = int(ws[row][3].value)
                hours = int(ws[row][4].value)

        subject = Subject(id, name, semester, hours, spec)

        return subject

@dataclass
class Group:
    name: str
    year: int
    spec: Specialization

    def __init__(self, new_name: str, new_year: int, new_spec: Specialization):
        self.name = new_name
        self.year = new_year
        self.spec = new_spec

@dataclass
class Exam:
    subject: Subject
    examDate: date
    year: str
    lecturer_fio: str

    def __init__(self, new_subject: Subject, new_examDate: date, new_year: str, new_lectFio: str):
        self.subject = new_subject
        self.examDate = new_examDate
        self.year = new_year
        self.lecturer_fio = new_lectFio

@dataclass
class ExamPoints:
    student: Student
    Points: int
    examPoints: int

    def __init__(self, new_student: Student, new_Points: int, new_examPoints: int):
        self.student = new_student
        self.Points = new_Points
        self.examPoints = new_examPoints

@dataclass
class Institute:
    students: List[Student]
    groups: List[Group]
    subjects: List[Subject]
    specs: List[Specialization]
    exams: List[Exam]
    exam_results: List[ExamPoints]

    def __init__(self):
        self.students: List[Student] = list()
        self.groups: List[Group] = list()
        self.subjects: List[Subject] = list()
        self.specs: List[Specialization] = list()
        self.exams: List[Exam] = list()
        self.exam_results: List[ExamPoints] = list()

    def addStudent(self, new_student: Student):
        if (isinstance(new_student, Student)) and (new_student not in self.students):
            if not new_student.fio == True:
                self.students.append(new_student)
                self.students.sort(key=lambda x: x.fio)
        else:
            raise Exception("Error")

    def addGroup(self, new_group: Group):
        if (isinstance(new_group, Group)) and (new_group not in self.groups):
            if not new_group.name == True:
                self.groups.append(new_group)
                self.groups.sort(key=lambda x: x.name)
        else:
            raise Exception("Error")

    def addSubject(self, new_subject: Subject):
        if (isinstance(new_subject, Subject)) and (new_subject not in self.subjects):
            if not new_subject.name == True:
                self.subjects.append(new_subject)
                self.subjects.sort(key=lambda x: x.name)
        else:
            raise Exception("Error")

    def addSpecs(self, new_spec: Specialization):
        if (isinstance(new_spec, Specialization)) and (new_spec not in self.specs):
            if not new_spec.name == True:
                self.specs.append(new_spec)
                self.specs.sort(key=lambda x: x.name)
        else:
            raise Exception("Error")

    def addExam(self, new_exam: Exam):
        if (isinstance(new_exam, Exam)) and (new_exam not in self.exams):
            self.exams.append(new_exam)
            self.exams.sort(key=lambda x: x.subject.name)
        else:
            raise Exception("Error")

    def addExamResult(self, new_examResult: ExamPoints):
        if (isinstance(new_examResult, ExamPoints)) and (new_examResult not in self.exam_results):
            self.exam_results.append(new_examResult)
            self.exam_results.sort(key=lambda x: x.examPoints)
        else:
            raise Exception("Error")

    def getSubject(self, compare_id: str):
        desired_sub = None

        for subject in self.subjects:
            if subject.id == compare_id:
                desired_sub = subject

        if desired_sub != None:
            return desired_sub
        else:
            raise Exception("Takoy subjecta net v spiske")

    def getStudent(self, compare_id: int):
        desired_stud = None

        for student in self.students:
            if student.id == compare_id:
                desired_stud = student

        if desired_stud != None:
            return desired_stud
        else:
            raise Exception("Takogo studenta net v spiske")

    def getExam(self, compare_name: str, exDate: date):
        desired_exam = None

        for exam in self.exams:
            if (exam.subject.name == compare_name) and (exam.examDate == exDate):
                desired_exam = exam

        if desired_exam != None:
            return desired_exam
        else:
            raise Exception("Takogo examena net v spiske")

    def getSpec(self, compare_name: str):
        desired_spec = None

        for spec in self.specs:
            if spec.name == compare_name:
                desired_spec = spec

        if desired_spec != None:
            return desired_spec
        else:
            raise Exception("Takoy specializacii net v spiske")

    def getGroup(self, compare_name: str):
        desired_group = None

        for group in self.groups:
            if group.name == compare_name:
                desired_group = group

        if desired_group != None:
            return desired_group
        else:
            raise Exception("Takoy grouppi net v spiske")

    def getExamPoints(self, compare_id: int):
        desired_exPoints = None

        for exPoints in self.exam_results:
            if exPoints.student.id == compare_id:
                desired_exPoints = exPoints

        if desired_exPoints != None:
            return desired_exPoints
        else:
            raise Exception("Resultatov examena etogo studenta net v spiske")

    def importSubjects(self, path):
        subject_wb = load_workbook(filename=path)
        ws = subject_wb.active

        for row in range(2, ws.max_row):
            id = str(ws[row][0].value)
            name = str(ws[row][1].value)
            spec = Specialization(ws[row][2].value)
            semester = int(ws[row][3].value)
            hours = int(ws[row][4].value)
            self.subjects.append(Subject(id, name, semester, hours, spec))

    def importStudents(self, path):
        subject_wb = load_workbook(filename=path)
        ws = subject_wb.active

        for row in range(2, ws.max_row + 1):
            fio = str(ws[row][0].value)
            id = int(ws[row][1].value)
            self.students.append(Student(fio, id))

    def importExams(self, path):
        subject_wb = load_workbook(filename=path)
        ws = subject_wb.active

        for row in range(2, ws.max_row):
            subject = Subject(ws[row][0].value)
            examDate = date(ws[row][2].value)
            lecturer_fio = str(ws[row][3].value)
            year = str(ws[row][4].value)
            self.subjects.append(Exam(subject, examDate, year, lecturer_fio))



inst = Institute()
path = 'C:/Users/Admin/Desktop/4.xlsx'
inst.importExams(path)
print(inst.exams)
#student1 = Student("Павлов Вячеслав", 175676)
#exPoints = ExamPoints(student1, 60, 28)
#student2 = Student("Афанасьев Вячеслав", 175676)
#exPoints2 = ExamPoints(student2, 60, 29)
#inst.addStudent(student1)
#inst.addExamResult(exPoints)
#inst.addStudent(student2)
#inst.addExamResult(exPoints2)
#print(inst.students)
#print(inst.exam_results)
